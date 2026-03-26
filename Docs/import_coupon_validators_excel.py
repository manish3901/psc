import argparse
import os
import re

from dotenv import load_dotenv
from openpyxl import load_workbook
from sqlalchemy import create_engine, text

load_dotenv()


def normalize_mobile(value: str) -> str:
    raw = "" if value is None else str(value)
    digits = re.sub(r"\D", "", raw)
    if len(digits) < 10:
        return ""
    return digits[-10:]


def header_map(headers):
    norm = []
    for h in headers:
        s = "" if h is None else str(h).strip().lower()
        s = re.sub(r"\s+", " ", s)
        norm.append(s)

    def idx(*candidates):
        for c in candidates:
            c = c.lower()
            if c in norm:
                return norm.index(c)
        return None

    return {
        "mobile": idx("mobile", "mobile no", "mobile number", "phone", "phone no", "phone number", "contact", "contact no"),
        "name": idx("name", "full name", "member name", "customer name"),
        "city": idx("city", "location", "area", "zone", "area/zone", "area zone"),
        "status": idx("status"),
    }


def main():
    parser = argparse.ArgumentParser(description="Import coupon validation mobiles from Excel into coupon_validators.")
    parser.add_argument("excel_path", nargs="?", default="Coupen_member_list.xlsx", help="Path to .xlsx file")
    parser.add_argument("--sheet", default=None, help="Sheet name (defaults to active sheet)")
    parser.add_argument("--update-existing", action="store_true", help="Update name/city/status if mobile exists")
    args = parser.parse_args()

    db_url = os.getenv(
        "DATABASE_URL",
        "postgresql+psycopg2://postgres:password@localhost:5432/psc_db",
    )
    engine = create_engine(db_url)

    excel_path = args.excel_path
    if not os.path.exists(excel_path):
        raise SystemExit(f"Excel file not found: {excel_path}")

    try:
        wb = load_workbook(excel_path, data_only=True)
    except PermissionError:
        raise SystemExit(f"Permission denied reading '{excel_path}'. Close the Excel file and try again.")

    ws = wb[args.sheet] if args.sheet else wb.active

    # Find header row (default: first row)
    headers = [c.value for c in ws[1]]
    mapping = header_map(headers)
    if mapping["mobile"] is None:
        raise SystemExit(
            "Could not find a Mobile column in row 1. Expected headers like: Mobile / Mobile No / Phone."
        )

    def cell(row, col_idx):
        if col_idx is None:
            return None
        v = row[col_idx].value if col_idx < len(row) else None
        return v

    rows_seen = set()
    inserted = 0
    updated = 0
    skipped_invalid = 0
    skipped_existing = 0

    with engine.begin() as conn:
        # Ensure table exists
        exists = conn.execute(
            text(
                "SELECT 1 FROM information_schema.tables WHERE table_name='coupon_validators' LIMIT 1"
            )
        ).first()
        if not exists:
            raise SystemExit("Table 'coupon_validators' not found. Run: python update_coupon_validators_schema.py")

        for r in ws.iter_rows(min_row=2):
            mobile = normalize_mobile(cell(r, mapping["mobile"]))
            if not mobile:
                skipped_invalid += 1
                continue

            if mobile in rows_seen:
                continue
            rows_seen.add(mobile)

            name = cell(r, mapping["name"])
            city = cell(r, mapping["city"])
            status = cell(r, mapping["status"])
            status = (str(status).strip().title() if status is not None and str(status).strip() else "Active")
            if status not in ("Active", "Disabled"):
                status = "Active"

            name = (str(name).strip() if name is not None and str(name).strip() else None)
            city = (str(city).strip() if city is not None and str(city).strip() else None)

            if args.update_existing:
                res = conn.execute(
                    text(
                        """
                        INSERT INTO coupon_validators (mobile_no, name, city, status)
                        VALUES (:mobile_no, :name, :city, :status)
                        ON CONFLICT (mobile_no)
                        DO UPDATE SET name = EXCLUDED.name,
                                      city = EXCLUDED.city,
                                      status = EXCLUDED.status
                        RETURNING xmax = 0 AS inserted
                        """
                    ),
                    {"mobile_no": mobile, "name": name, "city": city, "status": status},
                ).first()
                if res and res[0]:
                    inserted += 1
                else:
                    updated += 1
            else:
                res = conn.execute(
                    text(
                        """
                        INSERT INTO coupon_validators (mobile_no, name, city, status)
                        VALUES (:mobile_no, :name, :city, :status)
                        ON CONFLICT (mobile_no) DO NOTHING
                        RETURNING id
                        """
                    ),
                    {"mobile_no": mobile, "name": name, "city": city, "status": status},
                ).first()
                if res:
                    inserted += 1
                else:
                    skipped_existing += 1

    print(f"Imported from: {excel_path} (sheet: {ws.title})")
    print(f"Inserted: {inserted}")
    if args.update_existing:
        print(f"Updated: {updated}")
    print(f"Skipped invalid mobile: {skipped_invalid}")
    print(f"Skipped existing: {skipped_existing}")


if __name__ == "__main__":
    main()

