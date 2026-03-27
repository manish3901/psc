import os
import secrets
import string
from pathlib import Path

from sqlalchemy import func, or_, select

from flask import Blueprint, flash, redirect, render_template, request, send_file, session, url_for
from werkzeug.utils import secure_filename

from . import db
from .models import (
    CouponCode,
    CouponDistributionSettings,
    CouponMaster,
    CouponName,
    CouponPrizeAudit,
    CouponUser,
    CouponValidator,
)

main = Blueprint("main", __name__)


@main.route("/")
def index():
    # Default landing: user flow (scan & win). Admin is at /admin.
    return redirect(url_for("main.coupon_entry"))


def normalize_mobile(value):
    raw = "" if value is None else str(value)
    digits = "".join(ch for ch in raw if ch.isdigit())
    return digits[-10:] if len(digits) >= 10 else digits


def generate_coupon_barcode_value(prefix="CN"):
    for _ in range(20):
        candidate = f"{prefix}{secrets.randbelow(10**10):010d}"
        if not CouponName.query.filter_by(barcode_value=candidate).first():
            return candidate
    return f"{prefix}{secrets.token_hex(6).upper()}"


def generate_coupon_code_value(prefix="CC", length=8):
    alphabet = string.ascii_uppercase + string.digits
    for _ in range(40):
        candidate = prefix + "".join(secrets.choice(alphabet) for _ in range(length))
        if not CouponCode.query.filter_by(code=candidate).first():
            return candidate
    return prefix + secrets.token_hex(6).upper()


def require_admin():
    return bool(session.get("is_admin"))


def normalize_coupon_code(value):
    value = (value or "").strip().upper()
    return "".join(ch for ch in value if ch.isalnum() or ch in ("-", "_"))


def prize_level_cap(total_scratches: int, level: int, settings=None):
    unlock_at = int(getattr(settings, "unlock_at", 1000) if settings is not None else 1000)
    window1_end = int(getattr(settings, "window1_end", 1500) if settings is not None else 1500)
    l1_w1 = int(getattr(settings, "level1_cap_w1", 1) if settings is not None else 1)
    l1_w2 = int(getattr(settings, "level1_cap_w2", 2) if settings is not None else 2)
    l2_w1 = int(getattr(settings, "level2_cap_w1", 2) if settings is not None else 2)
    l3_w1 = int(getattr(settings, "level3_cap_w1", 2) if settings is not None else 2)

    # total_scratches is number of already revealed scratches (unique_code set) before selecting this reveal.
    if level == 1:
        if total_scratches < unlock_at:
            return 0
        if total_scratches < window1_end:
            return l1_w1
        return l1_w2
    if level in (2, 3):
        if total_scratches < unlock_at:
            return 0
        if total_scratches < window1_end:
            return l2_w1 if level == 2 else l3_w1
        return None
    return None


def prize_level_weight_multiplier(total_scratches: int, level: int, settings=None) -> float:
    unlock_at = int(getattr(settings, "unlock_at", 1000) if settings is not None else 1000)
    l1 = float(getattr(settings, "level1_mult", 0.35) if settings is not None else 0.35)
    l2 = float(getattr(settings, "level2_mult", 1.0) if settings is not None else 1.0)
    l3 = float(getattr(settings, "level3_mult", 1.8) if settings is not None else 1.8)
    l4 = float(getattr(settings, "level4_mult", 0.8) if settings is not None else 0.8)
    l5 = float(getattr(settings, "level5_mult", 1.0) if settings is not None else 1.0)
    l6 = float(getattr(settings, "level6_mult", 1.3) if settings is not None else 1.3)
    l7 = float(getattr(settings, "level7_mult", 1.7) if settings is not None else 1.7)

    if total_scratches < unlock_at and level in (1, 2, 3):
        return 0.0
    if level == 1:
        return max(0.0, l1)
    if level == 2:
        return max(0.0, l2)
    if level == 3:
        return max(0.0, l3)
    if level == 4:
        return max(0.0, l4)
    if level == 5:
        return max(0.0, l5)
    if level == 6:
        return max(0.0, l6)
    if level == 7:
        return max(0.0, l7)
    return 1.0


def generate_reference_code(length: int = 5) -> str:
    import random

    for _ in range(50):
        candidate = "".join(random.choices(string.digits, k=length))
        if not CouponUser.query.filter_by(unique_code=candidate).first():
            return candidate
    return "".join(secrets.choice(string.digits) for _ in range(length))


@main.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if require_admin():
        return redirect(url_for("main.admin_dashboard"))

    if request.method == "POST":
        password = request.form.get("password") or ""
        expected = os.getenv("ADMIN_PASSWORD") or ""
        if not expected:
            flash("ADMIN_PASSWORD is not set on the server.", "danger")
            return redirect(url_for("main.admin_login"))
        if secrets.compare_digest(password, expected):
            session["is_admin"] = True
            flash("Logged in.", "success")
            return redirect(url_for("main.admin_dashboard"))
        flash("Invalid password.", "danger")

    return render_template("admin/login.html")


@main.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    flash("Logged out.", "info")
    return redirect(url_for("main.admin_login"))


@main.route("/admin")
def admin_dashboard():
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    def _to_int(value):
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    coupon_names = CouponName.query.order_by(CouponName.coupon_name_id.desc()).all()

    m_sort = (request.args.get("m_sort") or "").strip().lower()
    m_dir = (request.args.get("m_dir") or "asc").strip().lower()
    if m_dir not in ("asc", "desc"):
        m_dir = "asc"

    coupon_masters_query = CouponMaster.query
    if m_sort == "level":
        if m_dir == "desc":
            coupon_masters_query = coupon_masters_query.order_by(
                CouponMaster.prize_level.is_(None),
                CouponMaster.prize_level.desc(),
                CouponMaster.coupon_master_id.desc(),
            )
        else:
            coupon_masters_query = coupon_masters_query.order_by(
                CouponMaster.prize_level.is_(None),
                CouponMaster.prize_level.asc(),
                CouponMaster.coupon_master_id.desc(),
            )
    else:
        coupon_masters_query = coupon_masters_query.order_by(CouponMaster.coupon_master_id.desc())

    coupon_masters = coupon_masters_query.all()

    edit_coupon_master = None
    edit_master_id = _to_int(request.args.get("edit_master_id"))
    if edit_master_id is not None:
        edit_coupon_master = CouponMaster.query.get(edit_master_id)

    # -----------------------------
    # Validators: filters/pagination
    # -----------------------------
    v_mobile = (request.args.get("v_mobile") or "").strip()
    v_name = (request.args.get("v_name") or "").strip()
    v_city = (request.args.get("v_city") or "").strip()
    v_status = (request.args.get("v_status") or "").strip().title()
    v_status = v_status if v_status in ("Active", "Disabled") else ""

    try:
        v_page = int(request.args.get("v_page") or 1)
    except ValueError:
        v_page = 1
    v_page = max(1, v_page)

    try:
        v_per_page = int(request.args.get("v_per_page") or 10)
    except ValueError:
        v_per_page = 10
    v_per_page = 10 if v_per_page not in (10, 25, 50, 100, 200) else v_per_page

    coupon_validators_query = CouponValidator.query
    if v_mobile:
        mobile_digits = normalize_mobile(v_mobile)
        if mobile_digits:
            coupon_validators_query = coupon_validators_query.filter(
                CouponValidator.mobile_no.ilike(f"%{mobile_digits}%")
            )
    if v_name:
        coupon_validators_query = coupon_validators_query.filter(
            CouponValidator.name.ilike(f"%{v_name}%")
        )
    if v_city:
        coupon_validators_query = coupon_validators_query.filter(
            CouponValidator.city.ilike(f"%{v_city}%")
        )
    if v_status:
        coupon_validators_query = coupon_validators_query.filter(CouponValidator.status == v_status)

    v_total = coupon_validators_query.order_by(None).count()
    v_total_pages = max(1, int((v_total + v_per_page - 1) / float(v_per_page)) if v_per_page else 1)
    if v_page > v_total_pages:
        v_page = v_total_pages

    coupon_validators = (
        coupon_validators_query.order_by(CouponValidator.id.desc())
        .offset((v_page - 1) * v_per_page)
        .limit(v_per_page)
        .all()
    )

    # -----------------------------
    # Users: filters/pagination/grouping
    # -----------------------------
    cu_name = (request.args.get("cu_name") or "").strip()
    cu_mobile = (request.args.get("cu_mobile") or "").strip()
    cu_coupon_name_id = (request.args.get("cu_coupon_name_id") or "").strip()
    cu_coupon_type = (request.args.get("cu_coupon_type") or "").strip()
    cu_coupon_code = (request.args.get("cu_coupon_code") or "").strip()
    cu_ref = (request.args.get("cu_ref") or "").strip()
    cu_group = (request.args.get("cu_group") or "").strip().lower()
    cu_group = cu_group if cu_group in ("mobile", "type", "code", "name") else ""

    try:
        cu_page = int(request.args.get("cu_page") or 1)
    except ValueError:
        cu_page = 1
    cu_page = max(1, cu_page)

    try:
        cu_per_page = int(request.args.get("cu_per_page") or 10)
    except ValueError:
        cu_per_page = 10
    cu_per_page = 10 if cu_per_page not in (10, 25, 50, 100, 200) else cu_per_page

    coupon_users_query = (
        CouponUser.query.outerjoin(CouponMaster, CouponUser.coupon_master_id == CouponMaster.coupon_master_id)
        .outerjoin(CouponName, CouponMaster.coupon_name_id == CouponName.coupon_name_id)
        .outerjoin(CouponCode, CouponUser.coupon_code_id == CouponCode.id)
    )

    if cu_name:
        like = f"%{cu_name}%"
        coupon_users_query = coupon_users_query.filter(
            or_(CouponUser.first_name.ilike(like), CouponUser.last_name.ilike(like))
        )
    if cu_mobile:
        mobile_digits = normalize_mobile(cu_mobile)
        if mobile_digits:
            coupon_users_query = coupon_users_query.filter(CouponUser.mobile_no.ilike(f"%{mobile_digits}%"))
    if cu_coupon_name_id:
        try:
            name_id_val = int(cu_coupon_name_id)
        except ValueError:
            name_id_val = None
        if name_id_val is not None:
            coupon_users_query = coupon_users_query.filter(CouponMaster.coupon_name_id == name_id_val)
    if cu_coupon_type:
        coupon_users_query = coupon_users_query.filter(CouponMaster.coupon_type == cu_coupon_type)
    if cu_coupon_code:
        coupon_users_query = coupon_users_query.filter(CouponCode.code.ilike(f"%{cu_coupon_code}%"))
    if cu_ref:
        coupon_users_query = coupon_users_query.filter(CouponUser.unique_code.ilike(f"%{cu_ref}%"))

    coupon_user_groups = []
    if cu_group:
        if cu_group == "mobile":
            group_col = CouponUser.mobile_no
            group_label = "Mobile"
        elif cu_group == "type":
            group_col = CouponMaster.coupon_type
            group_label = "Coupon Type"
        elif cu_group == "code":
            group_col = CouponCode.code
            group_label = "Coupon Code"
        else:
            group_col = CouponName.coupon_name
            group_label = "Coupon Name"

        grouped = (
            coupon_users_query.with_entities(
                group_col.label("group_value"),
                func.count(CouponUser.coupon_user_id).label("count"),
                func.min(CouponUser.created_at).label("first_at"),
                func.max(CouponUser.created_at).label("last_at"),
            )
            .filter(group_col.isnot(None))
            .group_by(group_col)
            .order_by(func.count(CouponUser.coupon_user_id).desc(), func.max(CouponUser.created_at).desc())
        )
        grouped_sub = grouped.subquery()
        cu_total = int(db.session.query(func.count()).select_from(grouped_sub).scalar() or 0)
        cu_total_pages = max(1, int((cu_total + cu_per_page - 1) / float(cu_per_page)) if cu_per_page else 1)
        if cu_page > cu_total_pages:
            cu_page = cu_total_pages

        coupon_user_groups = (
            db.session.query(grouped_sub)
            .offset((cu_page - 1) * cu_per_page)
            .limit(cu_per_page)
            .all()
        )
        coupon_users = []
    else:
        cu_total = coupon_users_query.order_by(None).count()
        cu_total_pages = max(1, int((cu_total + cu_per_page - 1) / float(cu_per_page)) if cu_per_page else 1)
        if cu_page > cu_total_pages:
            cu_page = cu_total_pages

        coupon_users = (
            coupon_users_query.order_by(CouponUser.coupon_user_id.desc())
            .offset((cu_page - 1) * cu_per_page)
            .limit(cu_per_page)
            .all()
        )

    coupon_type_options = [
        row[0]
        for row in db.session.query(CouponMaster.coupon_type)
        .distinct()
        .order_by(CouponMaster.coupon_type)
        .all()
        if row and row[0]
    ]
    coupon_codes = (
        CouponCode.query.order_by(CouponCode.id.desc()).limit(300).all()
        if hasattr(CouponCode, "query")
        else []
    )
    code_use_counts = {
        int(row[0]): int(row[1] or 0)
        for row in (
            db.session.query(CouponUser.coupon_code_id, func.count(CouponUser.coupon_user_id))
            .filter(CouponUser.coupon_code_id.isnot(None))
            .group_by(CouponUser.coupon_code_id)
            .all()
        )
        if row and row[0] is not None
    }

    edit_coupon_validator = None
    edit_validator_id = _to_int(request.args.get("edit_validator_id"))
    if edit_validator_id is not None:
        edit_coupon_validator = CouponValidator.query.get(edit_validator_id)

    dist_coupon_name_id = _to_int(request.args.get("dist_coupon_name_id"))
    if dist_coupon_name_id is None and coupon_names:
        dist_coupon_name_id = coupon_names[0].coupon_name_id
    dist_settings = (
        CouponDistributionSettings.query.get(dist_coupon_name_id) if dist_coupon_name_id is not None else None
    )

    dist_total_reveals = 0
    dist_level_stats = []
    if dist_coupon_name_id is not None:
        try:
            dist_total_reveals = int(
                (
                    db.session.query(func.count(CouponUser.coupon_user_id))
                    .join(CouponCode, CouponCode.id == CouponUser.coupon_code_id)
                    .filter(CouponCode.coupon_name_id == dist_coupon_name_id)
                    .filter(CouponUser.unique_code.isnot(None))
                    .scalar()
                    or 0
                )
            )
        except Exception:
            dist_total_reveals = 0

        masters_for_dist = (
            CouponMaster.query.filter_by(coupon_name_id=dist_coupon_name_id).all()
            if dist_coupon_name_id is not None
            else []
        )
        by_level = {}
        for m in masters_for_dist:
            lvl = int(m.prize_level) if m.prize_level is not None else None
            if lvl is None:
                continue
            r = by_level.setdefault(
                lvl,
                {
                    "level": lvl,
                    "prizes": 0,
                    "max_unlimited": False,
                    "max_total": 0,
                    "awarded_total": 0,
                    "remaining": None,
                },
            )
            r["prizes"] += 1
            max_allowed = int(getattr(m, "max_allowed", 0) or 0)
            if max_allowed <= 0:
                r["max_unlimited"] = True
            else:
                r["max_total"] += max_allowed
            r["awarded_total"] += int(getattr(m, "awarded_count", 0) or 0)

        for lvl in sorted(by_level.keys()):
            r = by_level[lvl]
            if r["max_unlimited"]:
                r["remaining"] = None
            else:
                r["remaining"] = max(0, int(r["max_total"]) - int(r["awarded_total"]))
            dist_level_stats.append(r)

    return render_template(
        "admin/dashboard.html",
        coupon_names=coupon_names,
        coupon_masters=coupon_masters,
        coupon_users=coupon_users,
        coupon_user_groups=coupon_user_groups,
        cu_name=cu_name,
        cu_mobile=cu_mobile,
        cu_coupon_name_id=cu_coupon_name_id,
        cu_coupon_type=cu_coupon_type,
        cu_coupon_code=cu_coupon_code,
        cu_ref=cu_ref,
        cu_group=cu_group,
        cu_page=cu_page,
        cu_per_page=cu_per_page,
        cu_total=cu_total,
        cu_total_pages=cu_total_pages,
        m_sort=m_sort,
        m_dir=m_dir,
        edit_coupon_master=edit_coupon_master,
        coupon_validators=coupon_validators,
        edit_coupon_validator=edit_coupon_validator,
        v_mobile=v_mobile,
        v_name=v_name,
        v_city=v_city,
        v_status=v_status,
        v_page=v_page,
        v_per_page=v_per_page,
        v_total=v_total,
        v_total_pages=v_total_pages,
        coupon_codes=coupon_codes,
        code_use_counts=code_use_counts,
        coupon_type_options=coupon_type_options,
        dist_coupon_name_id=dist_coupon_name_id,
        dist_settings=dist_settings,
        dist_total_reveals=dist_total_reveals,
        dist_level_stats=dist_level_stats,
    )


@main.route("/admin/coupon/name/create", methods=["POST"])
def coupon_name_create():
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    name = (request.form.get("coupon_name") or "").strip()
    description = (request.form.get("description") or "").strip() or None
    status = request.form.get("status", "Active")
    status = status if status in ("Active", "Disabled") else "Active"

    if not name:
        flash("Coupon name is required.", "danger")
        return redirect(url_for("main.admin_dashboard"))

    new_name = CouponName(
        coupon_name=name,
        description=description,
        status=status,
        barcode_value=generate_coupon_barcode_value(),
    )
    db.session.add(new_name)
    db.session.commit()
    flash(f"Coupon Name '{name}' created.", "success")
    return redirect(url_for("main.admin_dashboard"))


@main.route("/admin/coupon/name/update", methods=["POST"])
def coupon_name_update():
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    coupon_name_id = request.form.get("coupon_name_id")
    name = (request.form.get("coupon_name") or "").strip()
    description = (request.form.get("description") or "").strip() or None
    status = request.form.get("status", "Active")
    status = status if status in ("Active", "Disabled") else "Active"

    cn = CouponName.query.get(coupon_name_id)
    if not cn:
        flash("Coupon Name not found.", "danger")
        return redirect(url_for("main.admin_dashboard"))

    if not name:
        flash("Coupon name is required.", "danger")
        return redirect(url_for("main.admin_dashboard"))

    cn.coupon_name = name
    cn.description = description
    cn.status = status
    db.session.commit()
    flash("Coupon Name updated.", "success")
    return redirect(url_for("main.admin_dashboard"))


@main.route("/admin/coupon/code/generate", methods=["POST"])
def coupon_code_generate():
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    coupon_name_id = request.form.get("coupon_name_id")
    try:
        coupon_name_id = int(coupon_name_id)
    except (TypeError, ValueError):
        flash("Invalid coupon name.", "danger")
        return redirect(url_for("main.admin_dashboard", section="names"))

    cn = CouponName.query.get(coupon_name_id)
    if not cn:
        flash("Coupon name not found.", "danger")
        return redirect(url_for("main.admin_dashboard", section="names"))

    try:
        count = int(request.form.get("count") or 1)
    except ValueError:
        count = 1
    count = max(1, min(count, 500))

    created = 0
    for _ in range(count):
        code = generate_coupon_code_value(prefix="CC", length=8)
        db.session.add(CouponCode(coupon_name_id=cn.coupon_name_id, code=code, status="Active"))
        created += 1
    db.session.commit()
    flash(f"Generated {created} coupon codes for '{cn.coupon_name}'.", "success")
    return redirect(url_for("main.admin_dashboard", section="names"))


@main.route("/admin/coupon/code/update/<int:id>", methods=["POST"])
def coupon_code_update(id: int):
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    cc = CouponCode.query.get(id)
    if not cc:
        flash("Coupon code not found.", "danger")
        return redirect(url_for("main.admin_dashboard", section="names"))

    status = (request.form.get("status") or "Active").strip().title()
    status = status if status in ("Active", "Disabled") else "Active"
    cc.status = status
    db.session.commit()
    flash("Coupon code updated.", "success")

    ref = request.referrer or ""
    if ref.startswith(request.host_url):
        return redirect(ref)
    return redirect(url_for("main.admin_dashboard", section="names"))


@main.route("/admin/coupon/code/delete/<int:id>")
def coupon_code_delete(id: int):
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    cc = CouponCode.query.get(id)
    if not cc:
        flash("Coupon code not found.", "danger")
        return redirect(url_for("main.admin_dashboard", section="names"))

    used = (
        CouponUser.query.filter(CouponUser.coupon_code_id == id).order_by(CouponUser.coupon_user_id.desc()).first()
        is not None
    )
    if used:
        flash("Cannot delete: this coupon code is already used by users.", "danger")
        ref = request.referrer or ""
        if ref.startswith(request.host_url):
            return redirect(ref)
        return redirect(url_for("main.admin_dashboard", section="names"))

    db.session.delete(cc)
    db.session.commit()
    flash("Coupon code deleted.", "success")

    ref = request.referrer or ""
    if ref.startswith(request.host_url):
        return redirect(ref)
    return redirect(url_for("main.admin_dashboard", section="names"))


def _safe_filename(value: str, default: str = "qr") -> str:
    import re

    raw = (value or "").strip()
    if not raw:
        return default
    raw = re.sub(r"\s+", "-", raw)
    raw = re.sub(r"[^a-zA-Z0-9_.-]+", "", raw)
    raw = raw.strip("._-")
    return raw or default


@main.route("/admin/coupon-name/<int:coupon_name_id>/qr")
def coupon_name_qr(coupon_name_id: int):
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    cn = CouponName.query.get(coupon_name_id)
    if not cn:
        flash("Coupon name not found.", "danger")
        return redirect(url_for("main.admin_dashboard", section="names"))

    entry_url = url_for("main.coupon_entry", cn=cn.barcode_value, _external=True)

    try:
        import importlib

        qrcode = importlib.import_module("qrcode")
        SvgImage = importlib.import_module("qrcode.image.svg").SvgImage
    except Exception:
        flash("QR generation needs the 'qrcode' package. Install from requirements.txt.", "danger")
        return redirect(url_for("main.admin_dashboard", section="names"))

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=2,
    )
    qr.add_data(entry_url)
    qr.make(fit=True)
    img = qr.make_image(image_factory=SvgImage)

    from io import BytesIO

    buf = BytesIO()
    img.save(buf)
    buf.seek(0)

    download = (request.args.get("download") or "").strip().lower() in ("1", "true", "yes")
    filename = _safe_filename(cn.coupon_name, default=f"coupon_{cn.coupon_name_id}") + "_qr.svg"
    return send_file(
        buf,
        mimetype="image/svg+xml",
        as_attachment=download,
        download_name=filename,
        max_age=0,
    )


def _export_xlsx(filename: str, sheet_name: str, headers: list, rows: list, extra_sheets=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active

    def _fill_sheet(sheet, _sheet_name, _headers, _rows):
        sheet.title = (_sheet_name or "Sheet")[:31]
        sheet.append(list(_headers))
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for r in _rows:
            sheet.append(list(r))
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col[: min(len(col), 2000)]:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            sheet.column_dimensions[col_letter].width = min(50, max(10, max_len + 2))

    _fill_sheet(ws, sheet_name, headers, rows)

    if extra_sheets:
        for s in extra_sheets:
            if not s:
                continue
            ws2 = wb.create_sheet()
            _fill_sheet(ws2, s.get("sheet_name") or "Sheet", s.get("headers") or [], s.get("rows") or [])

    from io import BytesIO

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
        max_age=0,
    )


@main.route("/admin/export/validators")
def export_validators():
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    v_mobile = (request.args.get("v_mobile") or "").strip()
    v_name = (request.args.get("v_name") or "").strip()
    v_city = (request.args.get("v_city") or "").strip()
    v_status = (request.args.get("v_status") or "").strip().title()
    v_status = v_status if v_status in ("Active", "Disabled") else ""

    q = CouponValidator.query
    if v_mobile:
        mobile_digits = normalize_mobile(v_mobile)
        if mobile_digits:
            q = q.filter(CouponValidator.mobile_no.ilike(f"%{mobile_digits}%"))
    if v_name:
        q = q.filter(CouponValidator.name.ilike(f"%{v_name}%"))
    if v_city:
        q = q.filter(CouponValidator.city.ilike(f"%{v_city}%"))
    if v_status:
        q = q.filter(CouponValidator.status == v_status)

    data = []
    for v in q.order_by(CouponValidator.id.desc()).all():
        data.append(
            (
                v.id,
                v.mobile_no,
                v.name or "",
                v.city or "",
                v.status,
                v.created_at.strftime("%Y-%m-%d %H:%M:%S") if getattr(v, "created_at", None) else "",
            )
        )

    return _export_xlsx(
        filename="validators_export.xlsx",
        sheet_name="Validators",
        headers=["ID", "Mobile", "Name", "City", "Status", "Created At"],
        rows=data,
    )


@main.route("/admin/export/users")
def export_coupon_users():
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    cu_name = (request.args.get("cu_name") or "").strip()
    cu_mobile = (request.args.get("cu_mobile") or "").strip()
    cu_coupon_name_id = (request.args.get("cu_coupon_name_id") or "").strip()
    cu_coupon_type = (request.args.get("cu_coupon_type") or "").strip()
    cu_coupon_code = (request.args.get("cu_coupon_code") or "").strip()
    cu_ref = (request.args.get("cu_ref") or "").strip()
    cu_group = (request.args.get("cu_group") or "").strip().lower()
    cu_group = cu_group if cu_group in ("mobile", "type", "code", "name") else ""
    include_raw = (request.args.get("include_raw") or "").strip().lower() in ("1", "true", "yes")

    base = (
        CouponUser.query.outerjoin(CouponMaster, CouponUser.coupon_master_id == CouponMaster.coupon_master_id)
        .outerjoin(CouponName, CouponMaster.coupon_name_id == CouponName.coupon_name_id)
        .outerjoin(CouponCode, CouponUser.coupon_code_id == CouponCode.id)
    )

    if cu_name:
        like = f"%{cu_name}%"
        base = base.filter(or_(CouponUser.first_name.ilike(like), CouponUser.last_name.ilike(like)))
    if cu_mobile:
        mobile_digits = normalize_mobile(cu_mobile)
        if mobile_digits:
            base = base.filter(CouponUser.mobile_no.ilike(f"%{mobile_digits}%"))
    if cu_coupon_name_id:
        try:
            name_id_val = int(cu_coupon_name_id)
        except ValueError:
            name_id_val = None
        if name_id_val is not None:
            base = base.filter(CouponMaster.coupon_name_id == name_id_val)
    if cu_coupon_type:
        base = base.filter(CouponMaster.coupon_type == cu_coupon_type)
    if cu_coupon_code:
        base = base.filter(CouponCode.code.ilike(f"%{cu_coupon_code}%"))
    if cu_ref:
        base = base.filter(CouponUser.unique_code.ilike(f"%{cu_ref}%"))

    if cu_group:
        if cu_group == "mobile":
            group_col = CouponUser.mobile_no
            group_title = "Mobile"
        elif cu_group == "type":
            group_col = CouponMaster.coupon_type
            group_title = "Coupon Type"
        elif cu_group == "code":
            group_col = CouponCode.code
            group_title = "Coupon Code"
        else:
            group_col = CouponName.coupon_name
            group_title = "Coupon Name"

        grouped = (
            base.with_entities(
                group_col.label("group_value"),
                func.count(CouponUser.coupon_user_id).label("count"),
                func.min(CouponUser.created_at).label("first_at"),
                func.max(CouponUser.created_at).label("last_at"),
            )
            .filter(group_col.isnot(None))
            .group_by(group_col)
            .order_by(func.count(CouponUser.coupon_user_id).desc(), func.max(CouponUser.created_at).desc())
            .all()
        )

        rows = []
        for g in grouped:
            rows.append(
                (
                    g.group_value,
                    int(g.count or 0),
                    g.first_at.strftime("%Y-%m-%d %H:%M:%S") if g.first_at else "",
                    g.last_at.strftime("%Y-%m-%d %H:%M:%S") if g.last_at else "",
                )
            )

        extra = []
        if include_raw:
            all_users = base.order_by(CouponUser.coupon_user_id.desc()).all()
            audit_times = {}
            try:
                ids = [int(cu.coupon_user_id) for cu in all_users if cu and getattr(cu, "coupon_user_id", None)]
                if ids:
                    for uid, dt in (
                        db.session.query(CouponPrizeAudit.coupon_user_id, func.min(CouponPrizeAudit.created_at))
                        .filter(CouponPrizeAudit.coupon_user_id.in_(ids))
                        .group_by(CouponPrizeAudit.coupon_user_id)
                        .all()
                    ):
                        if uid is not None and dt is not None:
                            audit_times[int(uid)] = dt
            except Exception:
                audit_times = {}

            raw_rows = []
            for i, cu in enumerate(all_users, start=1):
                master = CouponMaster.query.get(cu.coupon_master_id) if cu.coupon_master_id else None
                cname = ""
                ctype = ""
                level = ""
                if master:
                    ctype = master.coupon_type or ""
                    level = master.prize_level if master.prize_level is not None else ""
                    cn = CouponName.query.get(master.coupon_name_id) if master.coupon_name_id else None
                    cname = cn.coupon_name if cn else ""

                if cu_group == "mobile":
                    gval = cu.mobile_no or ""
                elif cu_group == "type":
                    gval = ctype or ""
                elif cu_group == "code":
                    cc = CouponCode.query.get(cu.coupon_code_id) if cu.coupon_code_id else None
                    gval = cc.code if cc else ""
                else:
                    gval = cname or ""

                cc = CouponCode.query.get(cu.coupon_code_id) if cu.coupon_code_id else None

                scratched_at = audit_times.get(int(cu.coupon_user_id)) if getattr(cu, "coupon_user_id", None) else None
                scratched_at_str = scratched_at.strftime("%Y-%m-%d %H:%M:%S") if scratched_at else ""
                created_at_str = cu.created_at.strftime("%Y-%m-%d %H:%M:%S") if getattr(cu, "created_at", None) else ""

                raw_rows.append(
                    (
                        i,
                        cu.mobile_no or "",
                        gval,
                        created_at_str,
                        scratched_at_str,
                        cu.first_name or "",
                        cu.last_name or "",
                        cu.area_zone or "",
                        cname,
                        ctype,
                        (cc.code if cc else ""),
                        cu.unique_code or "",
                        level,
                    )
                )

            # For grouping by mobile, the group value is the mobile itself; drop the redundant Group column.
            if cu_group == "mobile":
                raw_headers = [
                    "Sr No",
                    "Mobile",
                    "Created At",
                    "Scratched At",
                    "First Name",
                    "Last Name",
                    "Area/Zone",
                    "Coupon Name",
                    "Coupon Type",
                    "Coupon Code",
                    "Reference Code",
                    "Prize Level",
                ]
                raw_rows = [
                    (
                        r[0],
                        r[1],
                        r[3],
                        r[4],
                        r[5],
                        r[6],
                        r[7],
                        r[8],
                        r[9],
                        r[10],
                        r[11],
                        r[12],
                    )
                    for r in raw_rows
                ]
            else:
                raw_headers = [
                    "Sr No",
                    "Mobile",
                    group_title,
                    "Created At",
                    "Scratched At",
                    "First Name",
                    "Last Name",
                    "Area/Zone",
                    "Coupon Name",
                    "Coupon Type",
                    "Coupon Code",
                    "Reference Code",
                    "Prize Level",
                ]

            extra.append(
                {
                    "sheet_name": "Raw Users",
                    "headers": raw_headers,
                    "rows": raw_rows,
                }
            )

        return _export_xlsx(
            filename="coupon_users_grouped_full_export.xlsx" if include_raw else "coupon_users_grouped_export.xlsx",
            sheet_name="Grouped Summary",
            headers=[group_title, "Count", "First At", "Last At"],
            rows=rows,
            extra_sheets=extra,
        )

    rows = []
    for i, cu in enumerate(base.order_by(CouponUser.coupon_user_id.desc()).all(), start=1):
        master = CouponMaster.query.get(cu.coupon_master_id) if cu.coupon_master_id else None
        cname = ""
        ctype = ""
        level = ""
        if master:
            ctype = master.coupon_type or ""
            level = master.prize_level if master.prize_level is not None else ""
            cn = CouponName.query.get(master.coupon_name_id) if master.coupon_name_id else None
            cname = cn.coupon_name if cn else ""

        cc = CouponCode.query.get(cu.coupon_code_id) if cu.coupon_code_id else None

        rows.append(
            (
                i,
                cu.created_at.strftime("%Y-%m-%d %H:%M:%S") if getattr(cu, "created_at", None) else "",
                cu.first_name or "",
                cu.last_name or "",
                cu.mobile_no or "",
                cu.area_zone or "",
                cname,
                ctype,
                (cc.code if cc else ""),
                cu.unique_code or "",
                level,
            )
        )

    return _export_xlsx(
        filename="coupon_users_export.xlsx",
        sheet_name="Coupon Users",
        headers=[
            "Sr No",
            "Created At",
            "First Name",
            "Last Name",
            "Mobile",
            "Area/Zone",
            "Coupon Name",
            "Coupon Type",
            "Coupon Code",
            "Reference Code",
            "Prize Level",
        ],
        rows=rows,
    )


@main.route("/admin/distribution/settings", methods=["POST"])
def distribution_settings_update():
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    coupon_name_id = request.form.get("coupon_name_id")
    try:
        coupon_name_id = int(coupon_name_id)
    except (TypeError, ValueError):
        flash("Invalid coupon name.", "danger")
        return redirect(url_for("main.admin_dashboard", section="distribution"))

    cn = CouponName.query.get(coupon_name_id)
    if not cn:
        flash("Coupon name not found.", "danger")
        return redirect(url_for("main.admin_dashboard", section="distribution"))

    def _int(name, default):
        val = request.form.get(name)
        if val is None or str(val).strip() == "":
            return default
        return int(val)

    def _float(name, default):
        val = request.form.get(name)
        if val is None or str(val).strip() == "":
            return default
        return float(val)

    try:
        unlock_at = max(0, _int("unlock_at", 1000))
        window1_end = max(unlock_at + 1, _int("window1_end", 1500))

        level1_cap_w1 = max(0, _int("level1_cap_w1", 1))
        level1_cap_w2 = max(level1_cap_w1, _int("level1_cap_w2", 2))
        level2_cap_w1 = max(0, _int("level2_cap_w1", 2))
        level3_cap_w1 = max(0, _int("level3_cap_w1", 2))

        level1_mult = max(0.0, _float("level1_mult", 0.35))
        level2_mult = max(0.0, _float("level2_mult", 1.0))
        level3_mult = max(0.0, _float("level3_mult", 1.8))
        level4_mult = max(0.0, _float("level4_mult", 0.8))
        level5_mult = max(0.0, _float("level5_mult", 1.0))
        level6_mult = max(0.0, _float("level6_mult", 1.3))
        level7_mult = max(0.0, _float("level7_mult", 1.7))
    except Exception:
        flash("Invalid values. Please enter numbers only.", "danger")
        return redirect(url_for("main.admin_dashboard", section="distribution", dist_coupon_name_id=coupon_name_id))

    settings = CouponDistributionSettings.query.get(coupon_name_id)
    if not settings:
        settings = CouponDistributionSettings(coupon_name_id=coupon_name_id)
        db.session.add(settings)

    settings.unlock_at = unlock_at
    settings.window1_end = window1_end
    settings.window2_end = max(window1_end + 1, int(getattr(settings, "window2_end", 2000) or 2000))
    settings.level1_cap_w1 = level1_cap_w1
    settings.level1_cap_w2 = level1_cap_w2
    settings.level2_cap_w1 = level2_cap_w1
    settings.level3_cap_w1 = level3_cap_w1
    settings.level1_mult = level1_mult
    settings.level2_mult = level2_mult
    settings.level3_mult = level3_mult
    settings.level4_mult = level4_mult
    settings.level5_mult = level5_mult
    settings.level6_mult = level6_mult
    settings.level7_mult = level7_mult

    db.session.commit()
    flash("Distribution settings updated.", "success")
    return redirect(url_for("main.admin_dashboard", section="distribution", dist_coupon_name_id=coupon_name_id))


@main.route("/admin/coupon/name/delete/<int:id>")
def coupon_name_delete(id):
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    cn = CouponName.query.get(id)
    if not cn:
        flash("Coupon Name not found.", "danger")
        return redirect(url_for("main.admin_dashboard"))

    if CouponMaster.query.filter_by(coupon_name_id=id).first():
        flash("Cannot delete: masters exist. Disable or delete masters first.", "danger")
        return redirect(url_for("main.admin_dashboard"))

    db.session.delete(cn)
    db.session.commit()
    flash("Coupon Name deleted.", "success")
    return redirect(url_for("main.admin_dashboard"))


def _save_prize_image(file_storage, name_id, coupon_type):
    if not file_storage or not file_storage.filename:
        return None

    allowed = {".png", ".jpg", ".jpeg", ".webp"}
    ext = Path(file_storage.filename).suffix.lower()
    if ext not in allowed:
        raise ValueError("Prize image must be PNG/JPG/WEBP.")

    upload_dir = Path(main.root_path) / "static" / "uploads" / "coupons"
    upload_dir.mkdir(parents=True, exist_ok=True)
    fname = secure_filename(file_storage.filename)
    fname = f"cm_{name_id}_{coupon_type}_{fname}".replace(" ", "_")
    file_storage.save(str(upload_dir / fname))
    return f"uploads/coupons/{fname}"


def _resolve_static_file(rel_path: str):
    try:
        static_root = (Path(main.root_path) / "static").resolve()
        rel = (rel_path or "").replace("\\", "/").lstrip("/")
        if rel.lower().startswith("static/"):
            rel = rel[7:]
        if not rel:
            return None
        candidate = (static_root / rel).resolve()
        if not str(candidate).startswith(str(static_root)):
            return None
        return candidate if candidate.is_file() else None
    except Exception:
        return None


@main.route("/admin/coupon/master/<int:id>/image")
def coupon_master_image(id: int):
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    cm = CouponMaster.query.get(id)
    if not cm:
        return "Coupon Master not found.", 404
    if not getattr(cm, "prize_image", None):
        return "No image uploaded for this prize.", 404

    path = _resolve_static_file(cm.prize_image)
    if not path:
        return "Image file not found on server. Please re-upload the image.", 404

    download = request.args.get("download") == "1"
    return send_file(str(path), as_attachment=download)


@main.route("/admin/coupon/master/create", methods=["POST"])
def coupon_master_create():
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    name_id = request.form.get("coupon_name_id")
    coupon_type = (request.form.get("coupon_type") or "").strip()
    max_allowed = request.form.get("max_allowed")
    prize_level = request.form.get("prize_level")
    weight = request.form.get("weight")
    status = request.form.get("status", "Active")
    status = status if status in ("Active", "Disabled") else "Active"

    if not name_id or not coupon_type:
        flash("Coupon Name and Coupon Type are required.", "danger")
        return redirect(url_for("main.admin_dashboard"))

    try:
        max_allowed_val = (
            int(max_allowed) if max_allowed is not None and str(max_allowed).strip() != "" else 0
        )
        prize_level_val = (
            int(prize_level) if prize_level is not None and str(prize_level).strip() != "" else None
        )
        weight_val = int(weight) if weight is not None and str(weight).strip() != "" else 1
    except ValueError:
        flash("Max Allowed / Prize Level / Weight must be numbers.", "danger")
        return redirect(url_for("main.admin_dashboard"))

    prize_image_path = None
    try:
        prize_image_path = _save_prize_image(request.files.get("prize_image"), name_id, coupon_type)
    except ValueError as e:
        flash(str(e), "danger")
        return redirect(url_for("main.admin_dashboard"))

    new_master = CouponMaster(
        coupon_name_id=int(name_id),
        coupon_type=coupon_type,
        max_allowed=max_allowed_val,
        prize_level=prize_level_val,
        weight=max(0, weight_val),
        status=status,
        prize_image=prize_image_path,
    )
    db.session.add(new_master)
    db.session.commit()
    flash("Coupon Master created.", "success")
    ref = request.referrer or ""
    if ref.startswith(request.host_url):
        return redirect(ref)
    return redirect(url_for("main.admin_dashboard", section="masters"))


@main.route("/admin/coupon/master/update", methods=["POST"])
def coupon_master_update():
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    master_id = request.form.get("coupon_master_id")
    name_id = request.form.get("coupon_name_id")
    coupon_type = (request.form.get("coupon_type") or "").strip()
    max_allowed = request.form.get("max_allowed")
    prize_level = request.form.get("prize_level")
    weight = request.form.get("weight")
    status = request.form.get("status", "Active")
    status = status if status in ("Active", "Disabled") else "Active"

    cm = CouponMaster.query.get(master_id)
    if not cm:
        flash("Coupon Master not found.", "danger")
        return redirect(url_for("main.admin_dashboard"))

    try:
        max_allowed_val = (
            int(max_allowed) if max_allowed is not None and str(max_allowed).strip() != "" else 0
        )
        prize_level_val = (
            int(prize_level) if prize_level is not None and str(prize_level).strip() != "" else None
        )
        weight_val = int(weight) if weight is not None and str(weight).strip() != "" else (cm.weight or 1)
    except ValueError:
        flash("Max Allowed / Prize Level / Weight must be numbers.", "danger")
        return redirect(url_for("main.admin_dashboard"))

    cm.coupon_name_id = int(name_id)
    cm.coupon_type = coupon_type
    cm.max_allowed = max_allowed_val
    cm.prize_level = prize_level_val
    cm.weight = max(0, weight_val)
    cm.status = status

    try:
        new_path = _save_prize_image(request.files.get("prize_image"), name_id, coupon_type)
        if new_path:
            cm.prize_image = new_path
    except ValueError as e:
        flash(str(e), "danger")
        return redirect(url_for("main.admin_dashboard"))

    db.session.commit()
    flash("Coupon Master updated.", "success")
    ref = request.referrer or ""
    if ref.startswith(request.host_url):
        return redirect(ref)
    return redirect(url_for("main.admin_dashboard", section="masters"))


@main.route("/admin/coupon/master/delete/<int:id>")
def coupon_master_delete(id):
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    cm = CouponMaster.query.get(id)
    if not cm:
        flash("Coupon Master not found.", "danger")
        return redirect(url_for("main.admin_dashboard"))

    if CouponUser.query.filter_by(coupon_master_id=id).first():
        flash("Cannot delete: users exist. Disable it instead.", "danger")
        return redirect(url_for("main.admin_dashboard"))

    db.session.delete(cm)
    db.session.commit()
    flash("Coupon Master deleted.", "success")
    ref = request.referrer or ""
    if ref.startswith(request.host_url):
        return redirect(ref)
    return redirect(url_for("main.admin_dashboard", section="masters"))


@main.route("/admin/coupon/user/delete/<int:id>")
def coupon_user_delete(id):
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    cu = CouponUser.query.get(id)
    if not cu:
        flash("Coupon User entry not found.", "danger")
        return redirect(url_for("main.admin_dashboard"))

    # Delete dependent audit logs first (FK is non-nullable).
    try:
        CouponPrizeAudit.query.filter(CouponPrizeAudit.coupon_user_id == id).delete(
            synchronize_session=False
        )
        db.session.flush()
    except Exception:
        db.session.rollback()
        flash("Unable to delete prize audit logs for this user.", "danger")
        ref = request.referrer or ""
        if ref.startswith(request.host_url):
            return redirect(ref)
        return redirect(url_for("main.admin_dashboard", section="users"))

    # If this entry had a revealed prize, roll back the awarded counter for accurate remaining/caps.
    try:
        if cu.coupon_master_id and cu.unique_code:
            db.session.query(CouponMaster).filter(
                CouponMaster.coupon_master_id == cu.coupon_master_id,
                CouponMaster.awarded_count > 0,
            ).update({CouponMaster.awarded_count: CouponMaster.awarded_count - 1})
            db.session.flush()
    except Exception:
        db.session.rollback()
        flash("User deleted, but unable to adjust awarded counters.", "warning")

    db.session.delete(cu)
    db.session.commit()
    flash("Coupon User entry deleted.", "success")
    ref = request.referrer or ""
    if ref.startswith(request.host_url):
        return redirect(ref)
    return redirect(url_for("main.admin_dashboard", section="users"))


@main.route("/admin/coupon/validator/create", methods=["POST"])
def coupon_validator_create():
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    mobile_no = normalize_mobile(request.form.get("mobile_no"))
    name = (request.form.get("name") or "").strip() or None
    city = (request.form.get("city") or "").strip() or None
    status = request.form.get("status", "Active")
    status = status if status in ("Active", "Disabled") else "Active"

    if len(mobile_no) != 10:
        flash("Please enter a valid 10-digit mobile number.", "danger")
        ref = request.referrer or ""
        if ref.startswith(request.host_url):
            return redirect(ref)
        return redirect(url_for("main.admin_dashboard", section="validators"))

    if CouponValidator.query.filter_by(mobile_no=mobile_no).first():
        flash("This mobile is already in validation list.", "danger")
        ref = request.referrer or ""
        if ref.startswith(request.host_url):
            return redirect(ref)
        return redirect(url_for("main.admin_dashboard", section="validators"))

    db.session.add(CouponValidator(mobile_no=mobile_no, name=name, city=city, status=status))
    db.session.commit()
    flash("Validation mobile added.", "success")
    ref = request.referrer or ""
    if ref.startswith(request.host_url):
        return redirect(ref)
    return redirect(url_for("main.admin_dashboard", section="validators"))


@main.route("/admin/coupon/validator/update/<int:id>", methods=["POST"])
def coupon_validator_update(id):
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    cv = CouponValidator.query.get(id)
    if not cv:
        flash("Validation entry not found.", "danger")
        return redirect(url_for("main.admin_dashboard", section="validators"))

    mobile_no = normalize_mobile(request.form.get("mobile_no"))
    name = (request.form.get("name") or "").strip() or None
    city = (request.form.get("city") or "").strip() or None
    status = request.form.get("status", "Active")
    status = status if status in ("Active", "Disabled") else "Active"

    if len(mobile_no) != 10:
        flash("Please enter a valid 10-digit mobile number.", "danger")
        return redirect(url_for("main.admin_dashboard", section="validators", edit_validator_id=id))

    conflict = (
        CouponValidator.query.filter(CouponValidator.mobile_no == mobile_no, CouponValidator.id != id).first()
    )
    if conflict:
        flash("This mobile is already in validation list.", "danger")
        return redirect(url_for("main.admin_dashboard", section="validators", edit_validator_id=id))

    cv.mobile_no = mobile_no
    cv.name = name
    cv.city = city
    cv.status = status
    db.session.commit()

    flash("Validator updated.", "success")
    ref = request.referrer or ""
    if ref.startswith(request.host_url):
        return redirect(ref)
    return redirect(url_for("main.admin_dashboard", section="validators"))


@main.route("/admin/coupon/validator/delete/<int:id>")
def coupon_validator_delete(id):
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    cv = CouponValidator.query.get(id)
    if not cv:
        flash("Validation entry not found.", "danger")
        return redirect(url_for("main.admin_dashboard"))

    db.session.delete(cv)
    db.session.commit()
    flash("Validation entry deleted.", "success")
    ref = request.referrer or ""
    if ref.startswith(request.host_url):
        return redirect(ref)
    return redirect(url_for("main.admin_dashboard", section="validators"))


@main.route("/admin/coupon/validator/upload", methods=["POST"])
def coupon_validator_upload():
    if not require_admin():
        return redirect(url_for("main.admin_login"))

    wants_json = "application/json" in (request.headers.get("Accept") or "") or request.headers.get(
        "X-Requested-With"
    ) == "fetch"

    excel = request.files.get("excel_file")
    update_existing = request.form.get("update_existing") == "1"
    sheet_name = (request.form.get("sheet_name") or "").strip()

    if not excel or not excel.filename:
        if wants_json:
            return {"success": False, "message": "Please select an Excel file."}, 400
        flash("Please select an Excel file.", "danger")
        return redirect(url_for("main.admin_dashboard"))

    if not (excel.filename or "").lower().endswith(".xlsx"):
        if wants_json:
            return {"success": False, "message": "Only .xlsx files are supported."}, 400
        flash("Only .xlsx files are supported.", "danger")
        return redirect(url_for("main.admin_dashboard"))

    from io import BytesIO

    try:
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(excel.read()), data_only=True)
    except Exception:
        if wants_json:
            return {"success": False, "message": "Unable to read the Excel file."}, 400
        flash("Unable to read the Excel file.", "danger")
        return redirect(url_for("main.admin_dashboard"))

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            if wants_json:
                return {"success": False, "message": f"Sheet '{sheet_name}' not found."}, 400
            flash(f"Sheet '{sheet_name}' not found.", "danger")
            return redirect(url_for("main.admin_dashboard"))
        ws = wb[sheet_name]
    else:
        ws = wb.active

    headers = [(c.value or "") for c in ws[1]]
    norm = []
    import re

    for h in headers:
        s = str(h).strip().lower()
        s = re.sub(r"\\s+", " ", s)
        norm.append(s)

    def idx(*candidates):
        for c in candidates:
            c = c.lower()
            if c in norm:
                return norm.index(c)
        return None

    mobile_idx = idx(
        "mobile",
        "mobile no",
        "mobile number",
        "phone",
        "phone no",
        "phone number",
        "contact",
        "contact no",
    )
    name_idx = idx("name", "full name", "member name", "customer name")
    city_idx = idx("city", "location", "area", "zone", "area/zone", "area zone")
    status_idx = idx("status")

    if mobile_idx is None:
        if wants_json:
            return {"success": False, "message": "Excel must have a Mobile column in the first row."}, 400
        flash("Excel must have a Mobile column in the first row.", "danger")
        return redirect(url_for("main.admin_dashboard"))

    inserted = 0
    updated = 0
    skipped_invalid = 0
    skipped_existing = 0
    skipped_duplicates = 0
    skipped_rows = []
    warning_rows = []
    seen = set()

    for idx_row, row in enumerate(ws.iter_rows(min_row=2), start=2):
        mobile_val = row[mobile_idx].value if mobile_idx < len(row) else None
        mobile_raw = "" if mobile_val is None else str(mobile_val)
        mobile = normalize_mobile(mobile_val)
        if len(mobile) != 10:
            skipped_invalid += 1
            skipped_rows.append(
                {"row": idx_row, "mobile_raw": mobile_raw, "mobile": mobile, "reason": "Invalid mobile"}
            )
            continue
        if mobile in seen:
            skipped_duplicates += 1
            skipped_rows.append(
                {"row": idx_row, "mobile_raw": mobile_raw, "mobile": mobile, "reason": "Duplicate in upload"}
            )
            continue
        seen.add(mobile)

        name = row[name_idx].value if name_idx is not None and name_idx < len(row) else None
        city = row[city_idx].value if city_idx is not None and city_idx < len(row) else None
        status = row[status_idx].value if status_idx is not None and status_idx < len(row) else None
        status = (str(status).strip().title() if status is not None and str(status).strip() else "Active")
        if status not in ("Active", "Disabled"):
            status = "Active"

        name = (str(name).strip() if name is not None and str(name).strip() else None)
        city = (str(city).strip() if city is not None and str(city).strip() else None)

        # Formatting warning if raw had extra characters but normalized fine.
        digits_raw = "".join(ch for ch in mobile_raw if ch.isdigit())
        if digits_raw and digits_raw != mobile:
            warning_rows.append(
                {
                    "row": idx_row,
                    "mobile_raw": mobile_raw,
                    "mobile": mobile,
                    "warning": "Normalized formatting",
                }
            )

        existing = CouponValidator.query.filter_by(mobile_no=mobile).first()
        if existing:
            if update_existing:
                existing.name = name
                existing.city = city
                existing.status = status
                updated += 1
            else:
                skipped_existing += 1
                skipped_rows.append(
                    {
                        "row": idx_row,
                        "mobile_raw": mobile_raw,
                        "mobile": mobile,
                        "reason": "Already exists",
                    }
                )
            continue

        db.session.add(CouponValidator(mobile_no=mobile, name=name, city=city, status=status))
        inserted += 1

    db.session.commit()
    message = (
        f"Import done (sheet: {ws.title}). Inserted: {inserted}, Updated: {updated}, "
        f"Skipped existing: {skipped_existing}, Duplicates: {skipped_duplicates}, Invalid: {skipped_invalid}."
    )

    if wants_json:
        return {
            "success": True,
            "message": message,
            "sheet": ws.title,
            "inserted": inserted,
            "updated": updated,
            "skipped_existing": skipped_existing,
            "skipped_duplicates": skipped_duplicates,
            "skipped_invalid": skipped_invalid,
            "skipped_rows": skipped_rows,
            "warning_rows": warning_rows,
        }

    flash(message, "success")
    return redirect(url_for("main.admin_dashboard", section="validators"))


@main.route("/coupon/validate-mobile", methods=["POST"])
def validate_mobile():
    data = request.get_json(silent=True) or {}
    mobile = normalize_mobile(data.get("mobile"))
    cv = CouponValidator.query.filter_by(mobile_no=mobile, status="Active").first()
    if cv:
        return {"found": True, "name": cv.name or "", "mobile": cv.mobile_no}
    return {"found": False}


@main.route("/coupon/validate-coupon-code", methods=["POST"])
def validate_coupon_code():
    data = request.get_json(silent=True) or {}
    cn_code = (data.get("cn_code") or data.get("cn") or "").strip()
    code = normalize_coupon_code(data.get("code"))
    if not code:
        return {"found": False, "message": "Please enter a coupon code."}, 400

    # If QR (coupon name) is not provided, infer the offer from the coupon code itself.
    # Codes are globally unique in this app, so this is safe and avoids user confusion.
    if not cn_code:
        cc = CouponCode.query.filter_by(code=code, status="Active").first()
        if not cc:
            return {"found": False, "message": "Coupon code not found."}
        cn = CouponName.query.get(cc.coupon_name_id) if getattr(cc, "coupon_name_id", None) else None
        if not cn or cn.status != "Active":
            return {"found": False, "message": "This offer is disabled. Please use a valid coupon."}
        return {"found": True, "coupon_name": cn.coupon_name, "code": cc.code, "cn_code": cn.barcode_value, "auto": True}

    cn = CouponName.query.filter_by(barcode_value=cn_code).first()
    if not cn or cn.status != "Active":
        return {"found": False, "message": "Invalid/disabled coupon QR. Please scan again."}

    cc = CouponCode.query.filter_by(code=code, coupon_name_id=cn.coupon_name_id, status="Active").first()
    if not cc:
        return {"found": False, "message": "Coupon code not found for this offer."}

    return {"found": True, "coupon_name": cn.coupon_name, "code": cc.code, "cn_code": cn.barcode_value}


@main.route("/coupon/entry")
def coupon_entry():
    cn_code = (request.args.get("cn") or "").strip()
    selected_coupon_name = None

    masters_query = CouponMaster.query.filter_by(status="Active")
    if cn_code:
        selected_coupon_name = CouponName.query.filter_by(barcode_value=cn_code).first()
        if selected_coupon_name and selected_coupon_name.status == "Active":
            masters_query = masters_query.filter_by(coupon_name_id=selected_coupon_name.coupon_name_id)
        else:
            selected_coupon_name = None
            masters_query = masters_query.filter_by(coupon_name_id=-1)

    active_coupons = masters_query.all()
    return render_template(
        "coupon/coupon_user_entry.html",
        coupons=active_coupons,
        selected_coupon_name=selected_coupon_name,
        cn_code=cn_code,
        company_profile=None,
        company_name=os.getenv("COMPANY_NAME", "Prakrutik SparshCare"),
    )


@main.route("/coupon/register", methods=["POST"])
def coupon_register():
    first_name = (request.form.get("first_name") or "").strip() or None
    last_name = (request.form.get("last_name") or "").strip() or None
    mobile_no = normalize_mobile(request.form.get("mobile_no"))
    area_zone = (request.form.get("area_zone") or "").strip() or None
    cn_code = (request.form.get("cn_code") or "").strip()
    coupon_code = normalize_coupon_code(request.form.get("coupon_code"))

    if not first_name or not last_name or not area_zone:
        return {"success": False, "message": "Please fill First Name, Last Name and Area/Zone."}

    if len(mobile_no) != 10:
        return {"success": False, "message": "Please enter a valid 10-digit mobile number."}

    cv = CouponValidator.query.filter_by(mobile_no=mobile_no, status="Active").first()
    if not cv:
        return {"success": False, "message": "Please try again with a registered mobile number."}

    if not coupon_code:
        return {"success": False, "message": "Coupon code is required."}

    cn = None
    cc = None
    if cn_code:
        cn = CouponName.query.filter_by(barcode_value=cn_code).first()
        if not cn or cn.status != "Active":
            return {"success": False, "message": "Invalid/disabled coupon QR. Please scan a valid coupon."}
        cc = CouponCode.query.filter_by(code=coupon_code, coupon_name_id=cn.coupon_name_id, status="Active").first()
        if not cc:
            return {"success": False, "message": "Coupon code not found for this offer."}
    else:
        # Fallback: infer coupon name from coupon code if QR isn't provided.
        cc = CouponCode.query.filter_by(code=coupon_code, status="Active").first()
        if not cc:
            return {"success": False, "message": "Coupon code not found for this offer."}
        cn = CouponName.query.get(cc.coupon_name_id) if getattr(cc, "coupon_name_id", None) else None
        if not cn or cn.status != "Active":
            return {"success": False, "message": "This offer is disabled. Please use a valid coupon."}

    existing_use = CouponUser.query.filter_by(mobile_no=mobile_no, coupon_code_id=cc.id).first()
    if existing_use:
        return {"success": False, "message": "This coupon code is already used by this mobile number."}

    new_user_coupon = CouponUser(
        first_name=first_name,
        last_name=last_name,
        mobile_no=mobile_no,
        area_zone=area_zone,
        coupon_master_id=None,
        coupon_code_id=cc.id,
        unique_code=None,
    )
    db.session.add(new_user_coupon)
    db.session.commit()

    return {"success": True, "coupon_user_id": new_user_coupon.coupon_user_id}


@main.route("/coupon/reveal-code", methods=["POST"])
def coupon_reveal_code():
    data = request.get_json(silent=True) or {}
    coupon_user_id = data.get("coupon_user_id")

    cu = CouponUser.query.get(coupon_user_id)
    if not cu:
        return {"success": False, "message": "Invalid coupon request."}, 400

    if cu.coupon_code_id is None:
        return {"success": False, "message": "Invalid coupon request."}, 400
    cc = CouponCode.query.get(cu.coupon_code_id)
    if not cc or cc.status != "Active":
        return {"success": False, "message": "Coupon code is not active."}, 400

    coupon_name_id = cc.coupon_name_id
    settings = CouponDistributionSettings.query.get(coupon_name_id)

    def _master_payload(master_row):
        prize_image_url = None
        if getattr(master_row, "prize_image", None):
            if _resolve_static_file(master_row.prize_image):
                prize_image_url = url_for("static", filename=master_row.prize_image)
        return {
            "success": True,
            "unique_code": cu.unique_code,
            "coupon_type": master_row.coupon_type,
            "prize_level": master_row.prize_level,
            "prize_image_url": prize_image_url,
        }

    if cu.unique_code:
        if cu.coupon_master_id:
            master = CouponMaster.query.get(cu.coupon_master_id)
            if master:
                return _master_payload(master)
        return {"success": True, "unique_code": cu.unique_code}

    total_scratches = (
        db.session.query(func.count(CouponUser.coupon_user_id))
        .join(CouponCode, CouponCode.id == CouponUser.coupon_code_id)
        .filter(CouponCode.coupon_name_id == coupon_name_id)
        .filter(CouponUser.unique_code.isnot(None))
        .scalar()
        or 0
    )

    masters = (
        CouponMaster.query.filter_by(coupon_name_id=coupon_name_id, status="Active")
        .order_by(CouponMaster.coupon_master_id.desc())
        .all()
    )
    if not masters:
        return {"success": False, "message": "No prizes configured for this offer."}, 400

    for _ in range(20):
        # Compute per-level awarded totals once per attempt (used for eligibility filtering).
        awarded_by_level = {
            int(row[0] or 0): int(row[1] or 0)
            for row in (
                db.session.query(CouponMaster.prize_level, func.coalesce(func.sum(CouponMaster.awarded_count), 0))
                .filter(CouponMaster.coupon_name_id == coupon_name_id)
                .group_by(CouponMaster.prize_level)
                .all()
            )
        }

        eligible = []
        weights = []
        for m in masters:
            if m.max_allowed and int(m.max_allowed) > 0 and int(getattr(m, "awarded_count", 0) or 0) >= int(
                m.max_allowed
            ):
                continue
            lvl = int(m.prize_level or 999)
            cap = prize_level_cap(int(total_scratches), lvl, settings=settings)
            if cap is not None and cap <= 0:
                continue
            # Cap is per *level total* (not per master).
            if cap is not None and int(awarded_by_level.get(lvl, 0)) >= int(cap):
                continue
            base_w = int(getattr(m, "weight", 1) or 1)
            base_w = max(0, base_w)
            mult = prize_level_weight_multiplier(int(total_scratches), lvl, settings=settings)
            eff = float(base_w) * float(mult)
            if eff <= 0:
                continue
            eligible.append(m)
            weights.append(eff)

        if not eligible:
            return {"success": False, "message": "All prizes are exhausted right now. Please try again later."}, 400

        import random

        chosen = random.choices(eligible, weights=weights, k=1)[0]
        chosen_level = int(chosen.prize_level or 999)
        chosen_cap = prize_level_cap(int(total_scratches), chosen_level, settings=settings)

        unique_code = generate_reference_code()
        try:
            # Atomic safety:
            # - Never exceed per-master max_allowed
            # - Never exceed per-level cap (levels 1..3 windows) even with concurrent reveals
            level_total_subq = (
                select(func.coalesce(func.sum(CouponMaster.awarded_count), 0))
                .where(CouponMaster.coupon_name_id == coupon_name_id)
                .where(CouponMaster.prize_level == chosen_level)
                .scalar_subquery()
            )

            q = CouponMaster.query.filter(CouponMaster.coupon_master_id == chosen.coupon_master_id).filter(
                or_(CouponMaster.max_allowed == 0, CouponMaster.awarded_count < CouponMaster.max_allowed)
            )
            if chosen_cap is not None:
                q = q.filter(level_total_subq < int(chosen_cap))

            updated = q.update({CouponMaster.awarded_count: CouponMaster.awarded_count + 1})
            if updated != 1:
                db.session.rollback()
                masters = [m for m in masters if m.coupon_master_id != chosen.coupon_master_id]
                continue

            cu.coupon_master_id = chosen.coupon_master_id
            cu.unique_code = unique_code
            db.session.add(
                CouponPrizeAudit(
                    coupon_user_id=cu.coupon_user_id,
                    coupon_name_id=coupon_name_id,
                    coupon_master_id=chosen.coupon_master_id,
                    prize_level=chosen.prize_level,
                    selection_mode="weighted",
                )
            )
            db.session.commit()
            return _master_payload(chosen)
        except Exception:
            db.session.rollback()
            continue

    return {"success": False, "message": "Unable to reveal right now. Please try again."}, 500

